#!/usr/bin/env python3
"""
Extract text content from .xlsx files for the STY Agent.
Outputs a plain-text representation of each sheet, suitable for Claude context.
Usage: python3 extract_xlsx.py <filepath> [max_rows_per_sheet]
"""

import sys
import openpyxl

MAX_ROWS = int(sys.argv[2]) if len(sys.argv) > 2 else 500
MAX_COLS = 20

def extract(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    output = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        output.append(f"## Sheet: {sheet_name}")
        rows_written = 0

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= MAX_ROWS:
                output.append(f"\n[Truncated: showing first {MAX_ROWS} rows]")
                break
            # Skip completely empty rows
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            cells = [str(cell) if cell is not None else "" for cell in row[:MAX_COLS]]
            # Strip trailing empty cells
            while cells and cells[-1] == "":
                cells.pop()
            if cells:
                output.append("\t".join(cells))
                rows_written += 1

        if rows_written == 0:
            output.append("(empty sheet)")

        output.append("")  # blank line between sheets

    wb.close()
    return "\n".join(output)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: extract_xlsx.py <filepath>", file=sys.stderr)
        sys.exit(1)
    try:
        print(extract(sys.argv[1]))
    except Exception as e:
        print(f"Error reading Excel file: {e}", file=sys.stderr)
        sys.exit(1)
